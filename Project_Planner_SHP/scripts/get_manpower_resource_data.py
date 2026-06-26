import os
import requests
import urllib3
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)

load_dotenv()
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# =============================================================================
# API CONFIGURATION
# =============================================================================

get_machine_api = (
    "https://api.sdprnel.com/bi/v1/resource/queries/"
    "00g75ugkse8UmN6xQ0i7__DAR7wvXDrQ/execute/"
    "?asset=DAR7wvXDrQ&asset_uid=DAR7wvXDrQ&organization=00g75ugkse8UmN6xQ0i7"
)

get_manpower_api = "https://api.sdprnel.com/bi/v1/resource/queries/00g75ugkse8UmN6xQ0i7__DAR7wvXDrQ/execute/?asset=DAR7wvXDrQ&asset_uid=DAR7wvXDrQ&organization=00g75ugkse8UmN6xQ0i7"

get_manhour_api = "https://api.sdprnel.com/bi/v1/resource/queries/00g75ugkse8UmN6xQ0i7__DAR7wvXDrQ/execute/?asset=DAR7wvXDrQ&asset_uid=DAR7wvXDrQ&organization=00g75ugkse8UmN6xQ0i7"

machine_payload = [
    {
        "table": "section_7_6Is2D0SMr_FieugCpEW",
        "filters": [
            {
                "field": "Forms ⇒ Sheet Pile Manpower and Machine ⇒ Shift Details ⇒ Date",
                "op": "relative_date_last",
                "value": ["-30_days", None, True],
                "meta": {
                    "is_stage_field": True,
                    "filter_config": {
                        "operator_value": ["-30_days", None, True],
                        "operator_name": "relative_date_last",
                        "anchor_period_time_amount": "30",
                        "anchor_period_starting_amount": None,
                        "anchor_period_starting_time_unit": None,
                        "anchor_period_starting_point": "today",
                        "anchor_period_starting_custom_date": None,
                        "anchor_period_time_unit": "days",
                        "anchor_period_include_period": True,
                    },
                },
            }
        ],
        "columns": [
            {
                "label": "Date",
                "field_type": "timestamp",
                "table_name": "forms_FieugCpEW",
                "field": "forms_FieugCpEW→field_-uzm4fSX0J",
                "alias": "Forms ⇒ Sheet Pile Manpower and Machine ⇒ Shift Details ⇒ Date",
            },
            {
                "label": "Machine",
                "field_type": "text",
                "table_name": "section_7_6Is2D0SMr_FieugCpEW",
                "field": "section_7_6Is2D0SMr_FieugCpEW→field_Q72eo_LRChv",
                "alias": "Sheet Pile Manpower and Machine ⇒ Machinery Status ⇒ Machine",
            },
            {
                "label": "Available",
                "field_type": "float",
                "table_name": "section_7_6Is2D0SMr_FieugCpEW",
                "field": "section_7_6Is2D0SMr_FieugCpEW→field_71l_vYOB0vY",
                "alias": "Sheet Pile Manpower and Machine ⇒ Machinery Status ⇒ Available",
            },
            {
                "label": "Required",
                "field_type": "float",
                "table_name": "section_7_6Is2D0SMr_FieugCpEW",
                "field": "section_7_6Is2D0SMr_FieugCpEW→field_-glV3uEsQDO",
                "alias": "Sheet Pile Manpower and Machine ⇒ Machinery Status ⇒ Required",
            },
        ],
        "joins": [
            {
                "table": {
                    "name": "forms_FieugCpEW",
                    "label": "Forms",
                    "alias": "forms_FieugCpEW",
                },
                "type": "inner",
                "on": [
                    {
                        "left": "section_7_6Is2D0SMr_FieugCpEW→Form ID",
                        "op": "=",
                        "right": "forms_FieugCpEW→ID",
                        "logic": "AND",
                        "right_label": "ID",
                        "right_table": "forms_FieugCpEW",
                        "left_label": "Form ID",
                        "left_table": "section_7_6Is2D0SMr_FieugCpEW",
                    }
                ],
            }
        ],
    }
]

manpower_payload = [{"table":"section_YbslHybA8bB_FieugCpEW","filters":[{"field":"Forms ⇒ Sheet Pile Manpower and Machine ⇒ Shift Details ⇒ Date","op":"relative_date_last","value":["-30_days",None,True],"meta":{"is_stage_field":True,"filter_config":{"operator_value":["-30_days",None,True],"operator_name":"relative_date_last","anchor_period_time_amount":"30","anchor_period_starting_amount":None,"anchor_period_starting_time_unit":None,"anchor_period_starting_point":"today","anchor_period_starting_custom_date":None,"anchor_period_time_unit":"days","anchor_period_include_period":True}}}],"columns":[{"label":"Date","field_type":"timestamp","table_name":"forms_FieugCpEW","field":"forms_FieugCpEW→field_-uzm4fSX0J","alias":"Forms ⇒ Sheet Pile Manpower and Machine ⇒ Shift Details ⇒ Date"},{"label":"Required","field_type":"float","table_name":"section_YbslHybA8bB_FieugCpEW","field":"section_YbslHybA8bB_FieugCpEW→field_Fzmur2itM-B","alias":"Sheet Pile Manpower and Machine ⇒ Manpower Status ⇒ Required"},{"label":"Available","field_type":"float","table_name":"section_YbslHybA8bB_FieugCpEW","field":"section_YbslHybA8bB_FieugCpEW→field_VRsvyVId0BC","alias":"Sheet Pile Manpower and Machine ⇒ Manpower Status ⇒ Available"}],"joins":[{"table":{"name":"forms_FieugCpEW","label":"Forms","alias":"forms_FieugCpEW"},"type":"inner","on":[{"left":"section_YbslHybA8bB_FieugCpEW→Form ID","op":"=","right":"forms_FieugCpEW→ID","logic":"AND","right_label":"ID","right_table":"forms_FieugCpEW","left_label":"Form ID","left_table":"section_YbslHybA8bB_FieugCpEW"}]}]}]

manhour_payload = [{"table":"section_YbslHybA8bB_FieugCpEW","filters":[{"field":"Date","op":"relative_date_last","value":["-30_days",None,True],"meta":{"is_stage_field":True,"filter_config":{"operator_value":["-30_days",None,True],"operator_name":"relative_date_last","anchor_period_time_amount":"30","anchor_period_starting_amount":None,"anchor_period_starting_time_unit":None,"anchor_period_starting_point":"today","anchor_period_starting_custom_date":None,"anchor_period_time_unit":"days","anchor_period_include_period":True}}}],"columns":[{"label":"Date","field_type":"timestamp","table_name":"forms_FieugCpEW","field":"forms_FieugCpEW→field_-uzm4fSX0J","alias":"Date"},{"label":"Actual \nWorkhours","field_type":"float","table_name":"forms_FieugCpEW","field":"forms_FieugCpEW→field_Klk5q915-ye","alias":"Work Hours"}],"joins":[{"table":{"name":"forms_FieugCpEW","label":"Forms","alias":"forms_FieugCpEW"},"type":"inner","on":[{"left":"section_YbslHybA8bB_FieugCpEW→Form ID","op":"=","right":"forms_FieugCpEW→ID","logic":"AND","right_label":"ID","right_table":"forms_FieugCpEW","left_label":"Form ID","left_table":"section_YbslHybA8bB_FieugCpEW"}]}]}]

BEARER_TOKEN = os.getenv("SDP_TOKEN", "")

HEADERS = {
    "Content-Type": "application/json",
    "Accept": "application/json",
}

if BEARER_TOKEN:
    HEADERS["Authorization"] = f"Bearer {BEARER_TOKEN}"

# =============================================================================
# EXCEL STYLING
# =============================================================================

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
ALT_FILL = PatternFill("solid", start_color="D6E4F0")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=10)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def thin_border():
    side = Side(style="thin", color="B0B0B0")
    return Border(left=side, right=side, top=side, bottom=side)


# =============================================================================
# API
# =============================================================================

def fetch_data(api_url, payload):
    print("Calling API...")

    response = requests.post(
        api_url,
        headers=HEADERS,
        json=payload,
        timeout=30,
        verify=False,
    )

    response.raise_for_status()

    return response.json()


# =============================================================================
# PARSE RESPONSE
# =============================================================================
def parse_rows(api_response):
    raw = api_response.get("data", [])

    if not raw:
        return []

    columns = api_response.get("metadata", {}).get("columns", [])
    keys = [c["name"] for c in columns]

    rows = []

    for item in raw:
        row = {}

        for key in keys:
            last_part = ""
            value = item.get(key)
            # Get the last segment of the alias
            if key.lower() in ['work hours', 'date']:
                last_part = key.lower()
            else:
                last_part = key.split("⇒")[-1].strip().lower() 

            if last_part == "" or last_part=="date":
                # Date column ends with "... ⇒"
                if isinstance(value, str):
                    try:
                        value = datetime.fromisoformat(
                            value.replace("Z", "+00:00")
                        ).strftime("%d-%b-%Y")
                    except ValueError:
                        pass

                row["date"] = value

            elif last_part.startswith("machin"):
                row["machine"] = value

            elif last_part.startswith("avail"):
                row["available"] = value

            elif last_part.startswith("work"):
                row["work hours"] = value
            # Ignore Required column
            elif last_part.startswith("requir"):
                continue

        rows.append(row)

    return rows
# =============================================================================
# WRITE EXCEL
# =============================================================================

def write_excel(machine_rows, manpower_rows, manhour_rows, output_file):
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    def write_sheet(ws, headers, rows):
        # ---------------- Header ----------------
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = thin_border()

        ws.row_dimensions[1].height = 22

        # ---------------- Data ----------------
        for row_idx, row in enumerate(rows, start=2):

            fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()

            for col_idx, value in enumerate(row.values(), start=1):

                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                cell.font = DATA_FONT
                cell.fill = fill
                cell.border = thin_border()

                if col_idx == 1:
                    cell.alignment = CENTER
                elif isinstance(value, (int, float)):
                    cell.alignment = RIGHT
                else:
                    cell.alignment = LEFT

        # Auto-fit columns
        for column_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            ws.column_dimensions[column_cells[0].column_letter].width = max(max_len + 3, 16)

        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # Machine Sheet
    # ------------------------------------------------------------------
    machine_ws = wb.create_sheet("Machine Status")

    write_sheet(
        machine_ws,
        headers=[
            "Shift Date",
            "Machine Type",
            "Machines Available",
        ],
        rows=machine_rows,
    )

    # ------------------------------------------------------------------
    # Manpower Sheet
    # ------------------------------------------------------------------
    manpower_ws = wb.create_sheet("Manpower Status")

    write_sheet(
        manpower_ws,
        headers=[
            "Shift Date",
            "Manpower Available",
        ],
        rows=manpower_rows,
    )

    # ------------------------------------------------------------------
    # Manhour Sheet
    # ------------------------------------------------------------------
    manhour_ws = wb.create_sheet("Manhour Status")

    write_sheet(
        manhour_ws,
        headers=[
            "Shift Date",
            "Manhour",
        ],
        rows=manhour_rows,
    )

    wb.save(output_file)

    print(f"Saved workbook -> {output_file}")
# =============================================================================
# MAIN
# =============================================================================
def main():

    try:
        machine_api_response = fetch_data(get_machine_api, machine_payload)
        manpower_api_response = fetch_data(get_manpower_api, manpower_payload)
        manhour_api_response = fetch_data(get_manhour_api, manhour_payload)
        
        print(
            f"[Machine] Rows returned : "
            f"{machine_api_response.get('metadata', {}).get('rowCount', 0)}"
        )
        
        print(
            f"[Manpower] Rows returned : "
            f"{manpower_api_response.get('metadata', {}).get('rowCount', 0)}"
        )

        print(
            f"[Manhour] Rows returned : "
            f"{manhour_api_response.get('metadata', {}).get('rowCount', 0)}"
        )

        machine_rows = parse_rows(machine_api_response)
        manpower_rows = parse_rows(manpower_api_response)
        manhour_rows = parse_rows(manhour_api_response)

        print(f"[Machine] Parsed rows  : {len(machine_rows)}")
        print(f"[Manpower] Parsed rows : {len(manpower_rows)}")
        print(f"[Manhour] Parsed rows : {len(manhour_rows)}")

    except requests.exceptions.HTTPError as e:

        if e.response.status_code == 401:
            print("Unauthorized. Please set SDP_TOKEN.")
            return

        raise

    except Exception as e:
        print(f"Error: {e}")
        raise

    write_excel(
        machine_rows,
        manpower_rows,
        manhour_rows,
        "./data/manpower_resources.xlsx",
    )


if __name__ == "__main__":
    main()